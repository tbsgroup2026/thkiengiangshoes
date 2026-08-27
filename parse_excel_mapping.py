#!/usr/bin/env python3
import openpyxl
import json
import sys
from pathlib import Path
from collections import Counter

def parse_excel_mapping():
    excel_path = Path(__file__).parent / "mapping_370_employees.xlsx"
    
    if not excel_path.exists():
        print(f"❌ File not found: {excel_path}")
        sys.exit(1)
    
    print(f"📂 Reading file: {excel_path}")
    print(f"📏 File size: {excel_path.stat().st_size:,} bytes")
    
    try:
        workbook = openpyxl.load_workbook(excel_path, data_only=True)
        worksheet = workbook.active
        
        print(f"\n📊 Worksheet: '{worksheet.title}'")
        print(f"📈 Total rows: {worksheet.max_row}")
        
        # Get header row (row 2 based on user's screenshot)
        headers = []
        for col_idx in range(1, worksheet.max_column + 1):
            cell_value = worksheet.cell(2, col_idx).value
            headers.append(str(cell_value).strip() if cell_value else "")
        
        print(f"\n📋 Headers ({len([h for h in headers if h])} columns):")
        for idx, h in enumerate(headers, 1):
            if h:
                print(f"  [{idx}] {h}")
        
        # Find key columns
        msnv_col = next((i for i, h in enumerate(headers, 1) if "MSNV" in h.upper()), None)
        name_col = next((i for i, h in enumerate(headers, 1) if "HỌ" in h.upper() and "TÊN" in h.upper()), None)
        dept_new_col = next((i for i, h in enumerate(headers, 1) if "BỘ PHẬN" in h.upper() and "NEW" in h.upper()), None)
        pb_new_col = next((i for i, h in enumerate(headers, 1) if "Phòng ban" in h and "NEW" in h.upper()), None)
        
        print(f"\n🔍 Key columns found:")
        print(f"  MSNV: column {msnv_col}")
        print(f"  Name: column {name_col}")
        print(f"  BỘ PHẬN (NEW): column {dept_new_col}")
        print(f"  Phòng ban (NEW): column {pb_new_col}")
        
        # Parse data rows
        mapping_data = []
        valid_count = 0
        invalid_rows = []
        
        for row_idx in range(3, worksheet.max_row + 1):  # Start from row 3 (skip headers)
            try:
                msnv_cell = worksheet.cell(row_idx, msnv_col) if msnv_col else None
                name_cell = worksheet.cell(row_idx, name_col) if name_col else None
                dept_new_cell = worksheet.cell(row_idx, dept_new_col) if dept_new_col else None
                pb_new_cell = worksheet.cell(row_idx, pb_new_col) if pb_new_col else None
                
                msnv = str(msnv_cell.value).strip() if msnv_cell and msnv_cell.value else ""
                name = str(name_cell.value).strip() if name_cell and name_cell.value else ""
                dept_new = str(dept_new_cell.value).strip() if dept_new_cell and dept_new_cell.value else ""
                pb_new = str(pb_new_cell.value).strip() if pb_new_cell and pb_new_cell.value else ""
                
                if msnv and (dept_new or pb_new):
                    mapping_data.append({
                        "msnv": msnv,
                        "name": name,
                        "department_new": dept_new if dept_new and dept_new != "None" else "",
                        "phong_ban_new": pb_new if pb_new and pb_new != "None" else "",
                        "row": row_idx
                    })
                    valid_count += 1
                elif msnv:
                    invalid_rows.append({
                        "row": row_idx,
                        "msnv": msnv,
                        "name": name,
                        "reason": "Missing department info"
                    })
            except Exception as e:
                invalid_rows.append({"row": row_idx, "error": str(e)})
        
        print(f"\n✅ Valid records: {valid_count}")
        print(f"❌ Invalid records: {len(invalid_rows)}")
        
        if invalid_rows:
            print(f"\n⚠️ Invalid rows:")
            for item in invalid_rows[:10]:  # Show first 10
                print(f"  Row {item.get('row')}: {item}")
        
        print(f"\n📊 Sample data (first 10):")
        for idx, item in enumerate(mapping_data[:10], 1):
            dept = item["department_new"] or item["phong_ban_new"]
            print(f"  [{idx}] {item['msnv']} | {item['name']} | {dept}")
        
        # Export to JSON
        json_path = Path(__file__).parent / "mapping_370_employees.json"
        with open(json_path, 'w', encoding='utf-8') as f:
            json.dump(mapping_data, f, ensure_ascii=False, indent=2)
        print(f"\n💾 Exported to: {json_path}")
        
        # Department statistics
        departments = set()
        for item in mapping_data:
            dept = item["department_new"] or item["phong_ban_new"]
            if dept:
                departments.add(dept)
        
        print(f"\n📈 Department groups found: {len(departments)}")
        dept_counter = Counter()
        for item in mapping_data:
            dept = item["department_new"] or item["phong_ban_new"]
            if dept:
                dept_counter[dept] += 1
        
        for dept, count in dept_counter.most_common():
            print(f"  • {dept}: {count} employees")
        
        # Target departments check
        target_depts = ["ĐH-QT", "NHÂN SỰ-HC", "KD PTSP", "QLCL & LAB", "CN-PPH & CI", "KHCB-TTPP"]
        print(f"\n🎯 Target 6 department groups:")
        for target in target_depts:
            count = dept_counter.get(target, 0)
            status = "✅" if count > 0 else "❌"
            print(f"  {status} {target}: {count} employees")
        
        print(f"\n✅ Parsing complete! Total valid: {valid_count}/370")
        
    except Exception as e:
        print(f"❌ Error: {e}", file=sys.stderr)
        import traceback
        traceback.print_exc()
        sys.exit(1)

if __name__ == "__main__":
    parse_excel_mapping()
