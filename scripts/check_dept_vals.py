import openpyxl
import json
from pathlib import Path

excel_path = Path("mapping_370_employees.xlsx")
wb = openpyxl.load_workbook(excel_path, data_only=True)
ws = wb.active

sample_rows = []
for row_idx in range(4, ws.max_row + 1):
    vals = [ws.cell(row_idx, c).value for c in range(1, ws.max_column + 1)]
    if not any(vals):
        continue
    msnv = str(vals[0]).strip() if vals[0] else ""
    name = str(vals[1]).strip() if vals[1] else ""
    pb_hien_tai = str(vals[4]).strip() if len(vals) > 4 and vals[4] is not None else ""
    pb_sap_xep = str(vals[7]).strip() if len(vals) > 7 and vals[7] is not None else ""
    sample_rows.append({"row": row_idx, "msnv": msnv, "name": name, "phong_ban_hien_tai": pb_hien_tai, "pb_sap_xep": pb_sap_xep})

with open("dept_check.json", "w", encoding="utf-8") as f:
    json.dump(sample_rows[:30], f, ensure_ascii=False, indent=2)

print("Saved dept_check.json")
