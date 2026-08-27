import openpyxl
import json

wb = openpyxl.load_workbook("mapping_370_employees.xlsx", data_only=True)
ws = wb.active

r2 = [str(ws.cell(2, c).value) for c in range(1, ws.max_column + 1)]
r3 = [str(ws.cell(3, c).value) for c in range(1, ws.max_column + 1)]
r4 = [str(ws.cell(4, c).value) for c in range(1, ws.max_column + 1)]

with open("test_rows.json", "w", encoding="utf-8") as f:
    json.dump({"row2": r2, "row3": r3, "row4": r4}, f, ensure_ascii=False, indent=2)

print("Wrote test_rows.json")
