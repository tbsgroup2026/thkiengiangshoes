import openpyxl
import json

wb = openpyxl.load_workbook("mapping_370_employees.xlsx", data_only=True)
ws = wb.active

row_vals = [ws.cell(373, c).value for c in range(1, ws.max_column + 1)]
headers = [ws.cell(2, c).value for c in range(1, ws.max_column + 1)]

with open("row373.json", "w", encoding="utf-8") as f:
    json.dump({"headers": headers, "row373": [str(v) for v in row_vals]}, f, ensure_ascii=False, indent=2)

print("Saved row373.json")
