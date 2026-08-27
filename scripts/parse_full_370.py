import openpyxl
import json
import sys
from pathlib import Path

sys.stdout.reconfigure(encoding='utf-8')

excel_path = Path("mapping_370_employees.xlsx")
wb = openpyxl.load_workbook(excel_path, data_only=True)
ws = wb.active

employees = []
for row_idx in range(4, ws.max_row + 1):
    row_vals = [ws.cell(row_idx, c).value for c in range(1, ws.max_column + 1)]
    if not any(row_vals):
        continue
    
    msnv = str(row_vals[0]).strip() if len(row_vals) > 0 and row_vals[0] is not None else ""
    name = str(row_vals[1]).strip() if len(row_vals) > 1 and row_vals[1] is not None else ""
    ngay_vao = str(row_vals[2]).strip() if len(row_vals) > 2 and row_vals[2] is not None else ""
    if " 00:00:00" in ngay_vao:
        ngay_vao = ngay_vao.split(" ")[0]
    
    vtcv_hien_tai = str(row_vals[3]).strip() if len(row_vals) > 3 and row_vals[3] is not None else ""
    phong_ban = str(row_vals[4]).strip() if len(row_vals) > 4 and row_vals[4] is not None else ""
    vtcv_sap = str(row_vals[5]).strip() if len(row_vals) > 5 and row_vals[5] is not None else ""
    vtcv_sap_xep = str(row_vals[6]).strip() if len(row_vals) > 6 and row_vals[6] is not None else ""
    pb_sap_xep = str(row_vals[7]).strip() if len(row_vals) > 7 and row_vals[7] is not None else ""
    bo_phan_new = str(row_vals[8]).strip() if len(row_vals) > 8 and row_vals[8] is not None else ""
    phong_ban_new = str(row_vals[9]).strip() if len(row_vals) > 9 and row_vals[9] is not None else ""
    ghi_chu = str(row_vals[10]).strip() if len(row_vals) > 10 and row_vals[10] is not None else ""

    if msnv and msnv != "None" and msnv != "STT" and msnv != "MSNV":
        role_code = "SUPER_ADMIN" if msnv in ["202608001", "202608002"] else ("TONG_GIAM_DOC" if vtcv_hien_tai == "TGĐ" else "CBCNV")
        
        employees.append({
            "id": f"emp_370_{len(employees)+1}",
            "empCode": msnv,
            "name": name if name != "None" else "N/A",
            "email": f"{msnv.lower()}@tbsgroup.vn",
            "phone": "0988 000 000",
            "title": vtcv_hien_tai if vtcv_hien_tai != "None" else "Cán Bộ Công Nhân Viên",
            "department": phong_ban if phong_ban != "None" else (bo_phan_new if bo_phan_new != "None" else "NHÂN SỰ-HC"),
            "roleCode": role_code,
            "status": "ACTIVE",
            "ngayVao": ngay_vao if ngay_vao != "None" else "-",
            "vtcvHienTai": vtcv_hien_tai if vtcv_hien_tai != "None" else "-",
            "phongBanHienTai": phong_ban if phong_ban != "None" else "-",
            "vtcvSap": vtcv_sap if vtcv_sap != "None" else "-",
            "vtcvSapXep": vtcv_sap_xep if vtcv_sap_xep != "None" else "-",
            "phongBanSapXep": pb_sap_xep if pb_sap_xep != "None" else "-",
            "boPhoanMoi": bo_phan_new if bo_phan_new != "None" else "-",
            "phongBanMoi": phong_ban_new if phong_ban_new != "None" else "-",
            "ghiChu": ghi_chu if ghi_chu != "None" else ""
        })

ts_content = """export interface EmployeeAccount {
  id: string;
  empCode: string;
  name: string;
  email: string;
  phone: string;
  title: string;
  department: string;
  roleCode: string;
  status: "ACTIVE" | "LOCKED";
  ngayVao?: string;
  vtcvHienTai?: string;
  phongBanHienTai?: string;
  vtcvSap?: string;
  vtcvSapXep?: string;
  phongBanSapXep?: string;
  boPhoanMoi?: string;
  phongBanMoi?: string;
  ghiChu?: string;
}

export const INITIAL_370_EMPLOYEES: EmployeeAccount[] = """ + json.dumps(employees, ensure_ascii=False, indent=2) + ";\n"

with open("web/src/lib/initialEmployees.ts", "w", encoding="utf-8") as f:
    f.write(ts_content)

print(f"DONE: Generated web/src/lib/initialEmployees.ts with {len(employees)} employee records.")
