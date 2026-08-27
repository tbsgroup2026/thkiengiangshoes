import openpyxl
import sys
from pathlib import Path

sys.stdout.reconfigure(encoding='utf-8')

excel_path = Path("mapping_370_employees.xlsx")
wb = openpyxl.load_workbook(excel_path, data_only=True)
ws = wb.active

sql_statements = [
    "-- SQL script to seed all 370 employees with full columns into D1 Database",
    "DELETE FROM users;"
]

count = 0
for row_idx in range(4, ws.max_row + 1):
    row_vals = [ws.cell(row_idx, c).value for c in range(1, ws.max_column + 1)]
    if not any(row_vals):
        continue
    
    msnv = str(row_vals[0]).strip() if len(row_vals) > 0 and row_vals[0] is not None else ""
    name = str(row_vals[1]).strip() if len(row_vals) > 1 and row_vals[1] is not None else ""
    ngay_vao = str(row_vals[2]).strip() if len(row_vals) > 2 and row_vals[2] is not None else ""
    if " 00:00:00" in ngay_vao:
        ngay_vao = ngay_vao.split(" ")[0]
    
    vtcv_hien_tai = str(row_vals[3]).strip() if len(row_vals) > 3 and row_vals[3] is not None else ""
    phong_ban = str(row_vals[4]).strip() if len(row_vals) > 4 and row_vals[4] is not None else ""
    vtcv_sap = str(row_vals[5]).strip() if len(row_vals) > 5 and row_vals[5] is not None else ""
    vtcv_sap_xep = str(row_vals[6]).strip() if len(row_vals) > 6 and row_vals[6] is not None else ""
    pb_sap_xep = str(row_vals[7]).strip() if len(row_vals) > 7 and row_vals[7] is not None else ""
    bo_phan_new = str(row_vals[8]).strip() if len(row_vals) > 8 and row_vals[8] is not None else ""
    phong_ban_new = str(row_vals[9]).strip() if len(row_vals) > 9 and row_vals[9] is not None else ""
    ghi_chu = str(row_vals[10]).strip() if len(row_vals) > 10 and row_vals[10] is not None else ""

    if msnv and msnv != "None" and msnv != "STT" and msnv != "MSNV":
        count += 1
        role_code = "SUPER_ADMIN" if msnv in ["202608001", "202608002"] else ("TONG_GIAM_DOC" if vtcv_hien_tai == "TGĐ" else "CBCNV")
        
        # Escape single quotes
        name_esc = name.replace("'", "''") if name != "None" else "N/A"
        title_esc = vtcv_hien_tai.replace("'", "''") if vtcv_hien_tai != "None" else "Cán Bộ Công Nhân Viên"
        dept_esc = phong_ban.replace("'", "''") if phong_ban != "None" else (bo_phan_new.replace("'", "''") if bo_phan_new != "None" else "NHÂN SỰ-HC")
        ngay_vao_esc = ngay_vao if ngay_vao != "None" else "2026-08-01"
        vtcv_hien_tai_esc = vtcv_hien_tai.replace("'", "''") if vtcv_hien_tai != "None" else "-"
        phong_ban_hien_tai_esc = phong_ban.replace("'", "''") if phong_ban != "None" else "-"
        vtcv_sap_esc = vtcv_sap.replace("'", "''") if vtcv_sap != "None" else "-"
        vtcv_sap_xep_esc = vtcv_sap_xep.replace("'", "''") if vtcv_sap_xep != "None" else "-"
        pb_sap_xep_esc = pb_sap_xep.replace("'", "''") if pb_sap_xep != "None" else "-"
        bo_phan_new_esc = bo_phan_new.replace("'", "''") if bo_phan_new != "None" else "-"

        sql = f"INSERT INTO users (emp_code, name, email, phone, title, department, role_code, password_hash, status, ngay_vao, vtcv_hien_tai, phong_ban_hien_tai, vtcv_sap, vtcv_sap_xep, pb_sap_xep, bo_phan_moi) VALUES ('{msnv}', '{name_esc}', '{msnv.lower()}@tbsgroup.vn', '0988 000 000', '{title_esc}', '{dept_esc}', '{role_code}', '123456', 'ACTIVE', '{ngay_vao_esc}', '{vtcv_hien_tai_esc}', '{phong_ban_hien_tai_esc}', '{vtcv_sap_esc}', '{vtcv_sap_xep_esc}', '{pb_sap_xep_esc}', '{bo_phan_new_esc}');"
        sql_statements.append(sql)

with open("web/seed_370_full_d1.sql", "w", encoding="utf-8") as f:
    f.write("\n".join(sql_statements) + "\n")

print(f"Generated web/seed_370_full_d1.sql with {count} INSERT statements.")
